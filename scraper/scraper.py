import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import requests

BASE_URL = "https://www.food.com/search/"
API_URL = "https://api.food.com/external/v1/nlp/search"

# Initialize Selenium WebDriver
options = Options()
options.add_argument("--headless")  # Run in background
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get(BASE_URL)

# # Function to load all recipes using infinite scroll
# def load_all_recipes():
#     # Get the total number of recipes from the search module title
#     total_recipes_text = driver.find_element(By.ID, 'searchModuleTitle').text
#     total_recipes = int(total_recipes_text.split(' ')[0].replace(',', ''))

#     # Infinite scroll to load all recipes
#     while True:
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         time.sleep(2)  # Adjust sleep time if necessary
#         loaded_recipes = len(driver.find_elements(By.CLASS_NAME, 'search-results .fd-recipe'))
#         if loaded_recipes >= total_recipes:
#             break

def get_recipes(url, payload):
    try:
        response = requests.post(url, json=payload)
        response.raise_for_status()  # Raise an exception for HTTP errors
        data = response.json().get("response", {}).get("results", [])
        for result in data:
            record_url = result.get('record_url')
            if record_url:
                recipe = extract_recipe_details(record_url)
                recipe["name"] = result.get("main_title")
                recipe["description"] = result.get("main_description")
                recipe["rating"] = result.get("main_rating")
                recipe["category"] = result.get("primary_category_name")
                recipe["url"] = record_url
                save_recipe_to_excel(recipe)

    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
        return None
    except ValueError:
        print("Response content is not valid JSON")
        return None


def clean_text(text):
    cleanded = text.strip().replace("\n", "")
    return " ".join(cleanded.split())

def save_recipe_to_excel(recipe, file_name='recipes.xlsx'):
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        # Write header
        headers = ['Name', 'Description', 'Facts', 'Directions', 'Ingredients', 'Nutrition', 'Category', 'Rating', "URL"]
        sheet.append(headers)

    # Write recipe data
    row = [
        recipe.get('name', 'N/A'),
        recipe.get('description', 'N/A'),
        str(recipe.get('facts', {})),
        '\n'.join(recipe.get('directions', [])),
        '\n'.join(recipe.get('ingredients', [])),
        str(recipe.get('nutrition', {})),
        recipe.get('category', 'N/A'),
        recipe.get('rating', 'N/A'),
        recipe.get('url', 'N/A')
    ]
    sheet.append(row)
    workbook.save(file_name)

# Function to extract recipe details
def extract_recipe_details(url):
    # recipes = []
    # recipe_links = [elem.get_attribute('href') for elem in driver.find_elements(By.CSS_SELECTOR, '.search-results .details h2.title a')]

    # for link in recipe_links:
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[1])
    driver.get(url)
    driver.find_element(By.CSS_SELECTOR, "button.link.facts__nutrition").click()
    
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    recipe = {}
    
    # # Extract recipe name
    # name_elem = soup.find('h1', class_='svelte-1muv3s8')
    # recipe['name'] = name_elem.text if name_elem else 'N/A'

    # # Extract recipe description
    # desc_elem = soup.find('div', class_='recipe-description paragraph')
    # if desc_elem:
    #     inner_desc_elem = desc_elem.find('div', class_='text svelte-1aswkii truncated')
    #     recipe['description'] = clean_text(inner_desc_elem.text) if inner_desc_elem else 'N/A'
    # else:
    #     recipe['description'] = 'N/A'
    
    # Extract facts
    facts = {}
    facts_elem = soup.find('div', class_='facts')
    if facts_elem:
        for item in facts_elem.find_all('div', class_='facts__item'):
            dt = item.find('dt').text
            dd = clean_text(item.find('dd').text)
            facts[dt] = dd
    recipe['facts'] = facts
    
    # Extract directions
    directions = []
    directions_elem = soup.find('section', class_='directions')
    if directions_elem:
        for li in directions_elem.find_all('li'):
            directions.append(clean_text(li.text))
    recipe['directions'] = directions
    
    # Extract ingredients
    ingredients = []
    ingredients_elem = soup.find('section', class_='ingredients')
    if ingredients_elem:
        for li in ingredients_elem.find_all('li'):
            ingredients.append(clean_text(li.text))
    recipe['ingredients'] = ingredients
    
    # Extract nutrition info
    nutrition = {}
    modal_elem = soup.find('div', class_='modal')
    if modal_elem:
        for p in modal_elem.find_all('p', class_='svelte-epeb0m'):
            bold_elem = p.find('span', class_='svelte-epeb0m')
            if bold_elem:
                key = bold_elem.text.replace(':', '').strip()
                value = p.text.replace(bold_elem.text, '').strip()
                nutrition[clean_text(key)] = clean_text(value)
        recipe['nutrition'] = nutrition
    
    # recipes.append(recipe)
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    return recipe

# Load all recipes
# load_all_recipes()

# Extract recipe details
# extract_recipe_details()

total_recipes_text = driver.find_element(By.ID, 'searchModuleTitle').text
total_recipes = int(total_recipes_text.split(' ')[0].replace(',', ''))
total_pages = int(total_recipes/10)
for i in range(1, total_pages+1):
    payload = {"pn": i}
    get_recipes(API_URL, payload)

# Close the WebDriver
driver.quit()
